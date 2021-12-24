from pathlib import Path
from typing import Sequence, Iterable

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from config import headlines_known_not_to_contain_table

Span = tuple[int, int]

TableSpan = tuple[Span, Span]

PATH_PLOTTING = Path("data/plots/")


class WorkbookReader:
    def __init__(self, path: Path, worksheet_names: Iterable[str]):
        self.workbook = openpyxl.load_workbook(path)

        self.worksheet_readers: dict[str, WorksheetReader] = {
            worksheet_name: WorksheetReader(self.workbook[worksheet_name]) for worksheet_name in worksheet_names
        }
        print(f'''Found sheets: "{'", "'.join(self.worksheet_readers.keys())}"''')

        self.plot_data_frames()

    def plot_data_frames(self):
        for worksheet_name, worksheet_reader in self.worksheet_readers.items():
            path_plotting = PATH_PLOTTING / worksheet_name
            path_plotting.mkdir(exist_ok=True)

            for data_frame_name, df in worksheet_reader.data_frames.items():
                df.plot()
                output_file_name = data_frame_name.replace("/", "")  # Cannot contain slash
                plt.savefig(path_plotting / output_file_name)


class WorksheetReader:
    def __init__(self, worksheet: Worksheet):
        self.worksheet = worksheet
        self.data_frames = self.get_data_frames()

    @staticmethod
    def is_headline_cell(cell: Cell):
        font: Font = cell.font
        return (
            font.underline == "single"
            and cell.hyperlink is None
            and cell.value not in headlines_known_not_to_contain_table
        )

    def get_table_headlines(self) -> list[Cell]:
        first_column: Sequence[Cell] = next(self.worksheet.iter_cols())
        return [cell for cell in first_column if self.is_headline_cell(cell)]

    def get_table_last_col(self, table_upper_left_cell: Cell) -> int:
        row = table_upper_left_cell.row
        col_span_cells: Sequence[Cell] = [
            col[0] for col in self.worksheet.iter_cols(min_row=row, max_row=row) if col[0].value
        ]
        return col_span_cells[-1].column

    def get_table_lower_right_cell(self, table_upper_left_cell: Cell, table_last_col: int) -> Cell:
        # All tables start by the left margin
        table_col_start = 1

        for row in self.worksheet.iter_rows(
            min_col=table_col_start, max_col=table_col_start, min_row=table_upper_left_cell.row
        ):
            cell: Cell = row[0]
            if not cell.value:
                # Previous cell is the last row of the table
                return self.worksheet.cell(row=cell.row - table_col_start, column=table_last_col)

        # Assume that row ends when sheet ends if we do not find an empty cell
        return self.worksheet.cell(row=self.worksheet.max_row, column=table_last_col)

    @staticmethod
    def _get_table_span(upper_left_cell: Cell, lower_right_cell: Cell) -> TableSpan:
        return (upper_left_cell.row, lower_right_cell.row), (upper_left_cell.column, lower_right_cell.column)

    def get_table_span(self, table_upper_left_cell: Cell) -> TableSpan:
        table_last_col = self.get_table_last_col(table_upper_left_cell)
        table_lower_right_cell = self.get_table_lower_right_cell(table_upper_left_cell, table_last_col=table_last_col)

        return self._get_table_span(upper_left_cell=table_upper_left_cell, lower_right_cell=table_lower_right_cell)

    def make_data_frame(self, table_span: TableSpan) -> pd.DataFrame:
        row_span, col_span = table_span

        data = [
            [
                cell if (cell := self.worksheet.cell(row=row, column=col).value) != "na" else np.nan
                for col in range(col_span[0], col_span[1] + 1)
            ]
            for row in range(row_span[0], row_span[1] + 1)
        ]
        header = data.pop(0)

        return pd.DataFrame(data, columns=header)

    def get_data_frame_from_headline(self, headline_cell: Cell) -> pd.DataFrame:
        call_below_headline_empty = not self.worksheet.cell(headline_cell.row + 1, headline_cell.column).value
        assert call_below_headline_empty

        table_upper_left_cell = self.worksheet.cell(row=headline_cell.row + 2, column=headline_cell.column)

        table_span = self.get_table_span(table_upper_left_cell)
        return self.make_data_frame(table_span=table_span)

    def get_data_frames(self) -> dict[str, pd.DataFrame]:
        table_headlines = self.get_table_headlines()
        return {
            headline_cell.value: self.get_data_frame_from_headline(headline_cell) for headline_cell in table_headlines
        }
