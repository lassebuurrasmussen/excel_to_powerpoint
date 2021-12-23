from pathlib import Path
from typing import Sequence

import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
from openpyxl.cell import Cell
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from config import file_name, sheet_name

Span = tuple[int, int]

TableSpan = tuple[Span, Span]


def is_headline_cell(cell: Cell):
    font: Font = cell.font
    return font.underline == "single" and cell.hyperlink is None


def _get_table_span(upper_left_cell: Cell, lower_right_cell: Cell) -> TableSpan:
    return (upper_left_cell.row, lower_right_cell.row), (upper_left_cell.column, lower_right_cell.column)


class WorksheetReader:
    def __init__(self, worksheet_name: str, path: Path):
        wb = openpyxl.load_workbook(path)
        self.worksheet: Worksheet = wb[worksheet_name]
        self.first_column: Sequence[Cell] = next(self.worksheet.iter_cols())
        self.data_frames = self.get_data_frames()

    def get_table_headlines(self) -> list[Cell]:
        return [cell for cell in self.first_column if is_headline_cell(cell)]

    def get_table_last_col(self, table_upper_left_cell: Cell) -> int:
        row = table_upper_left_cell.row
        col_span_cells: Sequence[Cell] = [
            col[0] for col in self.worksheet.iter_cols(min_row=row, max_row=row) if col[0].value
        ]
        return col_span_cells[-1].column

    def get_table_lower_right_cell(self, table_upper_left_cell: Cell, table_last_col: int) -> Cell:
        # All tables start by the left margin
        table_col_start = 1

        for row in self.worksheet.iter_rows(
            min_col=table_col_start, max_col=table_col_start, min_row=table_upper_left_cell.row
        ):
            cell: Cell = row[0]
            if not cell.value:
                # Previous cell is the last row of the table
                return self.worksheet.cell(row=cell.row - table_col_start, column=table_last_col)

        # Assume that row ends when sheet ends if we do not find an empty cell
        return self.worksheet.cell(row=self.worksheet.max_row, column=table_last_col)

    def get_table_span(self, table_upper_left_cell: Cell) -> TableSpan:
        table_last_col = self.get_table_last_col(table_upper_left_cell)

        table_lower_right_cell = self.get_table_lower_right_cell(table_upper_left_cell, table_last_col=table_last_col)

        return _get_table_span(upper_left_cell=table_upper_left_cell, lower_right_cell=table_lower_right_cell)

    def make_data_frame(self, table_span: TableSpan) -> pd.DataFrame:
        row_span, col_span = table_span

        data = [
            [self.worksheet.cell(row=row, column=col).value for col in range(col_span[0], col_span[1])]
            for row in range(row_span[0], row_span[1])
        ]
        header = data.pop(0)

        return pd.DataFrame(data, columns=header)

    def get_data_frame_from_headline(self, headline_cell: Cell) -> pd.DataFrame:
        call_below_headline_empty = not self.worksheet.cell(headline_cell.row + 1, headline_cell.column).value
        assert call_below_headline_empty

        table_upper_left_cell = self.worksheet.cell(row=headline_cell.row + 2, column=headline_cell.column)

        table_span = self.get_table_span(table_upper_left_cell)
        return self.make_data_frame(table_span=table_span)

    def get_data_frames(self) -> dict[str, pd.DataFrame]:
        table_headlines = self.get_table_headlines()
        return {
            headline_cell.value: self.get_data_frame_from_headline(headline_cell) for headline_cell in table_headlines
        }


if __name__ == "__main__":
    path_excel = Path("data") / file_name
    cpu_worksheet_reader = WorksheetReader(sheet_name, path_excel)
    table_names_string = "\n".join(cpu_worksheet_reader.data_frames.keys())
    print(f'Found the following tables on worksheet "{sheet_name}":\n\n{table_names_string}')
    for df_name, df in cpu_worksheet_reader.data_frames.items():
        df.plot(title=df_name)
        plt.savefig(f"data/plots/{df_name}")
